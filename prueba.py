import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

archivo_excel = load_workbook('Datos.xlsx')
hoja = archivo_excel.active
datos = []

for fila in hoja.iter_rows(values_only=True):
    datos.append(fila)

for x in datos:
    print (x)

# Inicializa el navegador
driver = webdriver.Chrome()

# Abre la página web
driver.get("https://adminnew.dinercol.co/")

# Encuentra el campo de entrada del formulario y lo llena con un valor
input_field = driver.find_element(By.NAME, "username")
input_field.send_keys("Dinercol")

input_field = driver.find_element(By.NAME, "password")
input_field.send_keys("Colombia1234*")

driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)

time.sleep(2)

error = 0
cargado = 0
contador = 0


current_url = driver.current_url
print("La URL actual es:", current_url)

time.sleep(20)
current_url = driver.current_url


if current_url == "https://adminnew.dinercol.co/#/layout/dashboard":

    driver.get("https://adminnew.dinercol.co/#/layout/overdue/overdueTicket")

    wait = WebDriverWait(driver, 10)  
    wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    for w in datos:

        contador=contador+1
        print("*** ", contador , " Bot en ejecucion Itenacion id_producto.", w , " ***")

        time.sleep(1)

        for i in range(11):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)

        active_element = driver.switch_to.active_element
        active_element.send_keys(w)

        active_element.send_keys(Keys.ENTER)

        for z in range(15):
                
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            time.sleep(1)
            active_ele = driver.switch_to.active_element
            active_ele.send_keys(Keys.ESCAPE)

        active_element2 = driver.switch_to.active_element
        active_element2.send_keys(Keys.ENTER)

        time.sleep(3)

        current_url_act = driver.current_url

        if current_url_act == 'https://adminnew.dinercol.co/#/layout/overdue/overdueTicket' :
            error = error + 1
            print('Id_producto no encontrado', w)
            driver.get("https://adminnew.dinercol.co/#/layout/overdue/overdueTicket")
            continue

        #Persona de contacto

        for i in range(37):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)

        active_element3 = driver.switch_to.active_element
        active_element3.send_keys(Keys.SPACE)

        time.sleep(1)

        #Contactos

        for i in range(1):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            time.sleep(1)


        active_element4 = driver.switch_to.active_element
        active_element4.send_keys(Keys.SPACE)

        time.sleep(1)

        #Tiempo de contacto

        for i in range(1):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            active_ele2 = driver.switch_to.active_element
            active_ele2.send_keys(Keys.ESCAPE)
            time.sleep(1)


        active_element5 = driver.switch_to.active_element
        active_element5.send_keys(Keys.SPACE)

        time.sleep(1)

        #Si estar en contacto

        for i in range(1):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            time.sleep(1)



        active_element6 = driver.switch_to.active_element
        active_element6.send_keys(Keys.ARROW_RIGHT)

        time.sleep(1)

        active_element7 = driver.switch_to.active_element
        active_element7.send_keys(Keys.SPACE)

        #Causas sin contestar

        for i in range(1):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            time.sleep(1)


        active_element8 = driver.switch_to.active_element
        active_element8.send_keys(Keys.SPACE)

        #Nota


        for i in range(1):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            time.sleep(1)


        active_element9 = driver.switch_to.active_element
        active_element9.send_keys("Gestion predictivo ")

        #Enviar

        for i in range(1):
            driver.find_element(By.TAG_NAME,"body").send_keys(Keys.TAB)
            time.sleep(1)


        active_element10 = driver.switch_to.active_element
        active_element10.send_keys(Keys.ENTER)
        
        time.sleep(1)

        wait = WebDriverWait(driver, 10)  # Configura un tiempo de espera máximo de 10 segundos
        wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

        current_url_act = driver.current_url

        if current_url_act == 'https://adminnew.dinercol.co/#/layout/overdue/overdueTicket' :
            cargado = cargado + 1
            print('Id_producto cargado con exito', w)

        driver.get("https://adminnew.dinercol.co/#/layout/overdue/overdueTicket")

        time.sleep(1)

    print('*** Ejecución Finalizada Con Exito *** ')
    print('✅ Correcto: ', cargado)
    print('❌ Errores: ', error)

elif current_url == "https://adminnew.dinercol.co/#/login?redirect=%23%2F":
    print("La URL actual es:", current_url)

